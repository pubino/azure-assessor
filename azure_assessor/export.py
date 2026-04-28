"""Export assessment results to Excel, CSV, and JSON."""

from __future__ import annotations

import csv
import json
from dataclasses import asdict
from io import StringIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from azure_assessor.models import AssessmentResult, ExportComponents


def _flatten_result(
    result: AssessmentResult, components: ExportComponents | None = None
) -> dict:
    """Flatten an AssessmentResult into a single-level dict for tabular export.

    Identifier fields (target_sku, region, timestamp) are always included so
    that filtered exports remain joinable.
    """
    components = components or ExportComponents.all()
    row: dict = {
        "target_sku": result.target_sku,
        "region": result.region,
        "timestamp": result.timestamp,
    }
    if components.summary:
        if result.availability:
            row["available"] = result.availability.available
            row["zones"] = ", ".join(result.availability.zones)
            row["restrictions"] = ", ".join(result.availability.restrictions)
        if result.quota:
            row["quota_family"] = result.quota.family
            row["quota_current"] = result.quota.current_usage
            row["quota_limit"] = result.quota.limit
            row["quota_available"] = result.quota.available
            row["quota_usage_pct"] = round(result.quota.usage_percent, 1)
        if result.pricing:
            row["price_hourly"] = result.pricing.retail_price
            row["price_currency"] = result.pricing.currency
            row["price_type"] = result.pricing.price_type
        if result.spot_pricing:
            row["spot_price_hourly"] = result.spot_pricing.retail_price
    if components.alternatives:
        row["num_alternatives"] = len(result.alternatives)
    if components.compatible_images:
        row["num_compatible_images"] = len(result.compatible_images)

    if components.cost_comparison and result.cost_comparison:
        vm_cost = next(
            (e.monthly_cost for e in result.cost_comparison if e.service_name == "Virtual Machines"),
            None,
        )
        cheapest = min(result.cost_comparison, key=lambda e: e.monthly_cost)
        row["cheapest_service"] = cheapest.service_name
        row["cheapest_tier"] = cheapest.tier
        row["cheapest_monthly"] = cheapest.monthly_cost
        if vm_cost and vm_cost > 0 and cheapest.service_name != "Virtual Machines":
            row["savings_vs_vm_pct"] = round(
                ((vm_cost - cheapest.monthly_cost) / vm_cost) * 100, 1
            )
        else:
            row["savings_vs_vm_pct"] = 0.0

    return row


def _filter_result_dict(
    data: dict, components: ExportComponents
) -> dict:
    """Strip top-level keys from a dataclasses.asdict() dump per components."""
    out = dict(data)
    if not components.summary:
        for k in ("availability", "quota", "pricing", "spot_pricing"):
            out.pop(k, None)
    if not components.alternatives:
        out.pop("alternatives", None)
    if not components.compatible_images:
        out.pop("compatible_images", None)
    if not components.cost_comparison:
        out.pop("cost_comparison", None)
    return out


def export_json(
    results: list[AssessmentResult],
    path: Path,
    components: ExportComponents | None = None,
) -> None:
    """Export results to JSON."""
    components = components or ExportComponents.all()
    data = [_filter_result_dict(asdict(r), components) for r in results]
    path.write_text(json.dumps(data, indent=2, default=str))


def export_json_string(
    results: list[AssessmentResult],
    components: ExportComponents | None = None,
) -> str:
    """Export results to a JSON string."""
    components = components or ExportComponents.all()
    data = [_filter_result_dict(asdict(r), components) for r in results]
    return json.dumps(data, indent=2, default=str)


def export_csv(
    results: list[AssessmentResult],
    path: Path,
    components: ExportComponents | None = None,
) -> None:
    """Export results to CSV."""
    if not results:
        path.write_text("")
        return
    components = components or ExportComponents.all()
    rows = [_flatten_result(r, components) for r in results]
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def export_csv_string(
    results: list[AssessmentResult],
    components: ExportComponents | None = None,
) -> str:
    """Export results to a CSV string."""
    if not results:
        return ""
    components = components or ExportComponents.all()
    rows = [_flatten_result(r, components) for r in results]
    fieldnames = list(rows[0].keys())
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue()


def export_excel(
    results: list[AssessmentResult],
    path: Path,
    components: ExportComponents | None = None,
) -> None:
    """Export results to Excel with formatting."""
    components = components or ExportComponents.all()
    wb = Workbook()
    # Drop the default sheet; we'll add only the sheets the caller selected.
    default = wb.active
    if default is not None:
        wb.remove(default)

    if components.summary:
        ws = wb.create_sheet("Summary")
        _write_summary_sheet(ws, results)

    if components.alternatives:
        ws_alt = wb.create_sheet("Alternatives")
        _write_alternatives_sheet(ws_alt, results)

    if components.compatible_images:
        ws_img = wb.create_sheet("Compatible Images")
        _write_images_sheet(ws_img, results)

    if components.cost_comparison:
        ws_cost = wb.create_sheet("Cost Comparison")
        _write_cost_comparison_sheet(ws_cost, results)

    if not wb.sheetnames:
        # openpyxl requires at least one sheet to save the workbook.
        wb.create_sheet("Empty")

    wb.save(str(path))


def _write_summary_sheet(ws, results: list[AssessmentResult]) -> None:
    """Write the summary sheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    headers = [
        "Target SKU", "Region", "Available", "Zones", "Restrictions",
        "Quota Family", "Usage", "Limit", "Available Quota", "Usage %",
        "Price/Hr", "Spot Price/Hr", "Currency", "Alternatives", "Compatible Images",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, result in enumerate(results, 2):
        flat = _flatten_result(result)
        ws.cell(row=row_idx, column=1, value=flat.get("target_sku", ""))
        ws.cell(row=row_idx, column=2, value=flat.get("region", ""))
        ws.cell(row=row_idx, column=3, value=flat.get("available", ""))
        ws.cell(row=row_idx, column=4, value=flat.get("zones", ""))
        ws.cell(row=row_idx, column=5, value=flat.get("restrictions", ""))
        ws.cell(row=row_idx, column=6, value=flat.get("quota_family", ""))
        ws.cell(row=row_idx, column=7, value=flat.get("quota_current", ""))
        ws.cell(row=row_idx, column=8, value=flat.get("quota_limit", ""))
        ws.cell(row=row_idx, column=9, value=flat.get("quota_available", ""))
        ws.cell(row=row_idx, column=10, value=flat.get("quota_usage_pct", ""))
        ws.cell(row=row_idx, column=11, value=flat.get("price_hourly", ""))
        ws.cell(row=row_idx, column=12, value=flat.get("spot_price_hourly", ""))
        ws.cell(row=row_idx, column=13, value=flat.get("price_currency", ""))
        ws.cell(row=row_idx, column=14, value=flat.get("num_alternatives", 0))
        ws.cell(row=row_idx, column=15, value=flat.get("num_compatible_images", 0))

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)


def _write_alternatives_sheet(ws, results: list[AssessmentResult]) -> None:
    """Write the alternatives sheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

    headers = [
        "Target SKU", "Region", "Alternative SKU", "Score",
        "vCPUs", "Memory (GB)", "Price/Hr", "Reasons",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for result in results:
        for alt in result.alternatives:
            ws.cell(row=row_idx, column=1, value=result.target_sku)
            ws.cell(row=row_idx, column=2, value=result.region)
            ws.cell(row=row_idx, column=3, value=alt.sku.name)
            ws.cell(row=row_idx, column=4, value=alt.compatibility_score)
            ws.cell(row=row_idx, column=5, value=alt.sku.vcpus)
            ws.cell(row=row_idx, column=6, value=alt.sku.memory_gb)
            ws.cell(row=row_idx, column=7, value=alt.price.retail_price if alt.price else "")
            ws.cell(row=row_idx, column=8, value="; ".join(alt.reasons))
            row_idx += 1


def _write_images_sheet(ws, results: list[AssessmentResult]) -> None:
    """Write the compatible images sheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

    headers = [
        "Target SKU", "Region", "Publisher", "Offer",
        "SKU", "Version", "OS Type", "Architecture", "HyperV Gen",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for result in results:
        for img in result.compatible_images:
            ws.cell(row=row_idx, column=1, value=result.target_sku)
            ws.cell(row=row_idx, column=2, value=result.region)
            ws.cell(row=row_idx, column=3, value=img.publisher)
            ws.cell(row=row_idx, column=4, value=img.offer)
            ws.cell(row=row_idx, column=5, value=img.sku)
            ws.cell(row=row_idx, column=6, value=img.version)
            ws.cell(row=row_idx, column=7, value=img.os_type)
            ws.cell(row=row_idx, column=8, value=img.architecture)
            ws.cell(row=row_idx, column=9, value=img.hyper_v_generation)
            row_idx += 1


def _write_cost_comparison_sheet(ws, results: list[AssessmentResult]) -> None:
    """Write the cost comparison sheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

    headers = [
        "Target SKU", "Region", "Service", "Tier/Plan",
        "vCPUs", "Memory (GB)", "Hourly Cost", "Monthly Cost",
        "Spot Monthly", "vs VM %", "Notes",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for result in results:
        vm_monthly = next(
            (e.monthly_cost for e in result.cost_comparison if e.service_name == "Virtual Machines"),
            None,
        )
        for est in result.cost_comparison:
            ws.cell(row=row_idx, column=1, value=result.target_sku)
            ws.cell(row=row_idx, column=2, value=result.region)
            ws.cell(row=row_idx, column=3, value=est.service_name)
            ws.cell(row=row_idx, column=4, value=est.tier)
            ws.cell(row=row_idx, column=5, value=est.vcpus if est.vcpus else "")
            ws.cell(row=row_idx, column=6, value=est.memory_gb if est.memory_gb else "")
            ws.cell(row=row_idx, column=7, value=est.hourly_cost)
            ws.cell(row=row_idx, column=8, value=est.monthly_cost)
            ws.cell(row=row_idx, column=9, value=est.spot_monthly if est.spot_monthly is not None else "")
            if vm_monthly and vm_monthly > 0 and est.service_name != "Virtual Machines":
                vs_vm = round(((est.monthly_cost - vm_monthly) / vm_monthly) * 100, 1)
                ws.cell(row=row_idx, column=10, value=vs_vm)
            else:
                ws.cell(row=row_idx, column=10, value="baseline" if est.service_name == "Virtual Machines" else "")
            ws.cell(row=row_idx, column=11, value="; ".join(est.notes))
            row_idx += 1

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
