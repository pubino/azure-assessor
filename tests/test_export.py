"""Tests for the export module."""

from __future__ import annotations

import csv
import json
from io import StringIO
from pathlib import Path

import pytest

from azure_assessor.export import (
    export_csv,
    export_csv_string,
    export_excel,
    export_json,
    export_json_string,
)
from azure_assessor.models import (
    AssessmentResult,
    ExportComponents,
    ImageInfo,
    PriceInfo,
    QuotaInfo,
    ServiceCostEstimate,
    SkuAvailability,
    SkuRecommendation,
    VmSku,
)


@pytest.fixture
def rich_result() -> AssessmentResult:
    """Assessment result with all fields populated."""
    return AssessmentResult(
        target_sku="Standard_D4s_v3",
        region="eastus",
        availability=SkuAvailability(
            sku_name="Standard_D4s_v3", region="eastus",
            zones=["1", "2", "3"], restrictions=[], available=True,
        ),
        quota=QuotaInfo(
            family="Standard DSv3 Family vCPUs", region="eastus",
            current_usage=24, limit=100,
        ),
        pricing=PriceInfo(
            sku_name="Standard_D4s_v3", region="eastus",
            retail_price=0.192, unit_price=0.192, currency="USD",
        ),
        spot_pricing=PriceInfo(
            sku_name="Standard_D4s_v3", region="eastus",
            retail_price=0.038, unit_price=0.038, is_spot=True,
        ),
        alternatives=[
            SkuRecommendation(
                sku=VmSku(
                    name="Standard_D8s_v3", family="standardDSv3Family",
                    size="D8s_v3", tier="Standard", vcpus=8, memory_gb=32.0,
                    max_data_disks=16, os_disk_size_gb=1023, max_nics=4,
                    accelerated_networking=True,
                ),
                compatibility_score=0.85,
                reasons=["Same family", "vCPUs: 8 (meets 4)"],
            ),
        ],
        compatible_images=[
            ImageInfo(
                publisher="Canonical",
                offer="0001-com-ubuntu-server-jammy",
                sku="22_04-lts",
                version="22.04.202401010",
            ),
        ],
        cost_comparison=[
            ServiceCostEstimate(
                service_name="Virtual Machines",
                tier="Standard_D4s_v3",
                monthly_cost=140.16,
                hourly_cost=0.192,
                vcpus=4,
                memory_gb=16.0,
                spot_monthly=27.74,
                notes=["Baseline VM cost"],
            ),
            ServiceCostEstimate(
                service_name="Container Apps",
                tier="Consumption",
                monthly_cost=105.12,
                hourly_cost=0.144,
                vcpus=4,
                memory_gb=16.0,
                notes=["vCPU rate: $0.000012/s"],
            ),
        ],
        timestamp="2026-04-02T12:00:00+00:00",
    )


class TestJsonExport:
    def test_export_json_string(self, rich_result):
        output = export_json_string([rich_result])
        data = json.loads(output)
        assert len(data) == 1
        assert data[0]["target_sku"] == "Standard_D4s_v3"
        assert data[0]["region"] == "eastus"
        assert data[0]["availability"]["available"] is True
        assert data[0]["pricing"]["retail_price"] == 0.192
        assert data[0]["spot_pricing"]["is_spot"] is True
        assert len(data[0]["alternatives"]) == 1
        assert len(data[0]["compatible_images"]) == 1

    def test_export_json_file(self, rich_result, tmp_path):
        path = tmp_path / "test.json"
        export_json([rich_result], path)
        assert path.exists()
        data = json.loads(path.read_text())
        assert len(data) == 1

    def test_export_empty(self, tmp_path):
        path = tmp_path / "empty.json"
        export_json([], path)
        data = json.loads(path.read_text())
        assert data == []


class TestCsvExport:
    def test_export_csv_string(self, rich_result):
        output = export_csv_string([rich_result])
        reader = csv.DictReader(StringIO(output))
        rows = list(reader)
        assert len(rows) == 1
        assert rows[0]["target_sku"] == "Standard_D4s_v3"
        assert rows[0]["available"] == "True"
        assert float(rows[0]["price_hourly"]) == 0.192
        assert float(rows[0]["spot_price_hourly"]) == 0.038

    def test_export_csv_file(self, rich_result, tmp_path):
        path = tmp_path / "test.csv"
        export_csv([rich_result], path)
        assert path.exists()
        content = path.read_text()
        assert "Standard_D4s_v3" in content

    def test_export_csv_empty(self, tmp_path):
        path = tmp_path / "empty.csv"
        export_csv([], path)
        assert path.read_text() == ""

    def test_csv_string_empty(self):
        assert export_csv_string([]) == ""


class TestExcelExport:
    def test_export_excel_file(self, rich_result, tmp_path):
        path = tmp_path / "test.xlsx"
        export_excel([rich_result], path)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_excel_has_sheets(self, rich_result, tmp_path):
        from openpyxl import load_workbook
        path = tmp_path / "test.xlsx"
        export_excel([rich_result], path)
        wb = load_workbook(str(path))
        assert "Summary" in wb.sheetnames
        assert "Alternatives" in wb.sheetnames
        assert "Compatible Images" in wb.sheetnames
        assert "Cost Comparison" in wb.sheetnames

    def test_excel_summary_data(self, rich_result, tmp_path):
        from openpyxl import load_workbook
        path = tmp_path / "test.xlsx"
        export_excel([rich_result], path)
        wb = load_workbook(str(path))
        ws = wb["Summary"]
        # Row 1 is header, Row 2 is data
        assert ws.cell(row=2, column=1).value == "Standard_D4s_v3"
        assert ws.cell(row=2, column=2).value == "eastus"
        assert ws.cell(row=2, column=3).value is True

    def test_excel_alternatives_data(self, rich_result, tmp_path):
        from openpyxl import load_workbook
        path = tmp_path / "test.xlsx"
        export_excel([rich_result], path)
        wb = load_workbook(str(path))
        ws = wb["Alternatives"]
        assert ws.cell(row=2, column=3).value == "Standard_D8s_v3"
        assert ws.cell(row=2, column=4).value == 0.85

    def test_excel_images_data(self, rich_result, tmp_path):
        from openpyxl import load_workbook
        path = tmp_path / "test.xlsx"
        export_excel([rich_result], path)
        wb = load_workbook(str(path))
        ws = wb["Compatible Images"]
        assert ws.cell(row=2, column=3).value == "Canonical"

    def test_excel_cost_comparison_data(self, rich_result, tmp_path):
        from openpyxl import load_workbook
        path = tmp_path / "test.xlsx"
        export_excel([rich_result], path)
        wb = load_workbook(str(path))
        ws = wb["Cost Comparison"]
        # Row 1 is header, Row 2 = VM, Row 3 = Container Apps
        assert ws.cell(row=2, column=3).value == "Virtual Machines"
        assert ws.cell(row=2, column=8).value == 140.16
        assert ws.cell(row=3, column=3).value == "Container Apps"
        assert ws.cell(row=3, column=8).value == 105.12

    def test_export_empty_excel(self, tmp_path):
        path = tmp_path / "empty.xlsx"
        export_excel([], path)
        assert path.exists()


class TestCostComparisonFlatten:
    def test_flatten_includes_cheapest(self, rich_result):
        from azure_assessor.export import _flatten_result
        flat = _flatten_result(rich_result)
        assert flat["cheapest_service"] == "Container Apps"
        assert flat["cheapest_monthly"] == 105.12
        assert flat["savings_vs_vm_pct"] > 0

    def test_flatten_no_cost_comparison(self):
        from azure_assessor.export import _flatten_result
        result = AssessmentResult(target_sku="Standard_D4s_v3", region="eastus")
        flat = _flatten_result(result)
        assert "cheapest_service" not in flat


class TestMultipleResults:
    def test_json_multiple(self, rich_result):
        results = [rich_result, rich_result]
        output = export_json_string(results)
        data = json.loads(output)
        assert len(data) == 2

    def test_csv_multiple(self, rich_result):
        results = [rich_result, rich_result]
        output = export_csv_string(results)
        reader = csv.DictReader(StringIO(output))
        rows = list(reader)
        assert len(rows) == 2


class TestSelectiveExport:
    """Selective export honors ExportComponents toggles across all formats."""

    def test_csv_summary_only(self, rich_result):
        comps = ExportComponents(
            summary=True, alternatives=False, compatible_images=False, cost_comparison=False,
        )
        output = export_csv_string([rich_result], comps)
        reader = csv.DictReader(StringIO(output))
        row = next(iter(reader))
        # Identifiers always present
        assert row["target_sku"] == "Standard_D4s_v3"
        # Summary fields present
        assert "available" in row
        assert "price_hourly" in row
        # Disabled component fields absent
        assert "num_alternatives" not in row
        assert "num_compatible_images" not in row
        assert "cheapest_service" not in row

    def test_csv_cost_only(self, rich_result):
        comps = ExportComponents(
            summary=False, alternatives=False, compatible_images=False, cost_comparison=True,
        )
        output = export_csv_string([rich_result], comps)
        reader = csv.DictReader(StringIO(output))
        row = next(iter(reader))
        assert row["target_sku"] == "Standard_D4s_v3"
        assert "cheapest_service" in row
        # Summary fields filtered out
        assert "available" not in row
        assert "price_hourly" not in row

    def test_csv_all_off_keeps_identifiers(self, rich_result):
        comps = ExportComponents.none()
        output = export_csv_string([rich_result], comps)
        reader = csv.DictReader(StringIO(output))
        row = next(iter(reader))
        assert set(row.keys()) == {"target_sku", "region", "timestamp"}

    def test_json_filters_top_level_keys(self, rich_result):
        comps = ExportComponents(
            summary=True, alternatives=False, compatible_images=False, cost_comparison=False,
        )
        data = json.loads(export_json_string([rich_result], comps))
        record = data[0]
        # Identifiers + summary fields stay
        assert record["target_sku"] == "Standard_D4s_v3"
        assert record.get("availability") is not None
        # Disabled lists/sections removed
        assert "alternatives" not in record
        assert "compatible_images" not in record
        assert "cost_comparison" not in record

    def test_json_cost_only_drops_summary(self, rich_result):
        comps = ExportComponents(
            summary=False, alternatives=False, compatible_images=False, cost_comparison=True,
        )
        data = json.loads(export_json_string([rich_result], comps))
        record = data[0]
        assert "availability" not in record
        assert "quota" not in record
        assert "pricing" not in record
        assert "spot_pricing" not in record
        assert "cost_comparison" in record

    def test_excel_only_selected_sheets(self, rich_result, tmp_path: Path):
        comps = ExportComponents(
            summary=True, alternatives=False, compatible_images=False, cost_comparison=True,
        )
        path = tmp_path / "filtered.xlsx"
        export_excel([rich_result], path, comps)

        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert "Summary" in wb.sheetnames
        assert "Cost Comparison" in wb.sheetnames
        assert "Alternatives" not in wb.sheetnames
        assert "Compatible Images" not in wb.sheetnames

    def test_excel_all_off_creates_empty_workbook(self, rich_result, tmp_path: Path):
        path = tmp_path / "empty.xlsx"
        export_excel([rich_result], path, ExportComponents.none())

        from openpyxl import load_workbook
        wb = load_workbook(path)
        # openpyxl requires at least one sheet
        assert len(wb.sheetnames) >= 1

    def test_default_components_match_legacy_behavior(self, rich_result):
        """Calling without components yields the same output as before."""
        legacy = export_csv_string([rich_result])
        explicit = export_csv_string([rich_result], ExportComponents.all())
        assert legacy == explicit
